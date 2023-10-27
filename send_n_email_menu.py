# Email của người gửi phải có tính năng: Allow less secure apps
# Tính năng này sẽ có khi email là của một tổ chức như email của trường
# Email bình thường abc@gmail.com không có tính năng này
# Đăng nhập email của trường và
# Vào đây https://myaccount.google.com/lesssecureapps để bật sang ON

import smtplib, ssl
import tkinter	as	tk
import tkinter.messagebox as mb
import	tkinter.filedialog	as	fd
import os
import openpyxl
from xls2xlsx import XLS2XLSX


class	App(tk.Tk):
    def	__init__(self):
        super().__init__()
        self.wb = None
        self.email = tk.StringVar()
        self.password = tk.StringVar()
        self.thong_tin = tk.StringVar()
        self.ds_email = None

        self.title('Gửi Điểm Cho Sinh Viên')
        menu = tk.Menu(self)
        file_menu = tk.Menu(menu, tearoff=0)

        file_menu.add_command(label="Open Excel File", command = self.onOpenExcel)
        file_menu.add_command(label="Send Email", command = self.onSendEmail)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command = self.destroy)

        help_menu = tk.Menu(menu, tearoff=0)
        help_menu.add_command(label="About", command = self.onAbout)

        menu.add_cascade(label="File", menu=file_menu)
        menu.add_cascade(label="Help", menu=help_menu)
        self.config(menu=menu)

        vscroll = tk.Scrollbar(self, orient = tk.VERTICAL)
        hscroll = tk.Scrollbar(self, orient = tk.HORIZONTAL)
        self.txt_bang_diem = tk.Text(self, font = ('Consolas', 13), width = 40, wrap = 'none',
                                yscrollcommand = vscroll.set, xscrollcommand = hscroll.set)
        vscroll.configure(command = self.txt_bang_diem.yview)
        hscroll.configure(command = self.txt_bang_diem.xview)
        self.txt_bang_diem.grid(row = 0, column = 0, padx = 10, pady = 10)
        vscroll.grid(row = 0, column = 1, padx = 10, pady = 10, sticky = tk.NS)
        hscroll.grid(row = 1, column = 0, padx = 10, pady = 10, sticky = tk.EW)

        frm_nguoi_gui = tk.LabelFrame(self)
        lbl_thong_tin = tk.Label(frm_nguoi_gui, text = 'Thông tin GV sẽ gắn cuối email:')
        ent_thong_tin = tk.Entry(frm_nguoi_gui, width = 32, font = ('Consolas', 12), textvariable = self.thong_tin) 
        lbl_email = tk.Label(frm_nguoi_gui, text = 'Email:',)
        ent_email = tk.Entry(frm_nguoi_gui, width = 32, font = ('Consolas', 12), textvariable = self.email)
        lbl_password = tk.Label(frm_nguoi_gui, text = 'Password:')
        ent_password = tk.Entry(frm_nguoi_gui, show = '*', font = ('Consolas', 12), textvariable = self.password)
        btn_send = tk.Button(frm_nguoi_gui, text = 'Send Email', width = 12, command = self.onSendEmail)

        self.vscroll_lst = tk.Scrollbar(frm_nguoi_gui, orient = tk.VERTICAL)
        self.lst_ket_qua = tk.Listbox(frm_nguoi_gui, font = ('Consolas', 8), height = 19, yscrollcommand = self.vscroll_lst.set)
        self.vscroll_lst.config(command=self.lst_ket_qua.yview)

        lbl_thong_tin.grid(row = 0, column = 0, padx = 10, pady = 5, sticky = tk.NW)
        ent_thong_tin.grid(row = 1, column = 0, padx = 10, pady = 0, sticky = tk.NW)
        lbl_email.grid(row = 2, column = 0, padx = 10, pady = 5, sticky = tk.NW)
        ent_email.grid(row = 3, column = 0, padx = 10, pady = 0, sticky = tk.NW)
        lbl_password.grid(row = 4, column = 0, padx = 10, pady = 5, sticky = tk.NW)
        ent_password.grid(row = 5, column = 0, padx = 10, pady = 5, sticky = tk.NW)
        btn_send.grid(row = 6, column = 0, padx = 10, pady = 5, sticky = tk.NW)
        self.lst_ket_qua.grid(row = 7, column = 0, padx = 10, pady = 5, sticky = tk.NSEW)
        self.vscroll_lst.grid(row = 7, column = 1, padx = 10, pady = 5, sticky = tk.NS)

        frm_nguoi_gui.grid(row = 0, column = 2, padx = 10, pady = 10, sticky = tk.NSEW)


    def get_danh_sach_sinh_vien(self):
        sheet = self.wb.active
        max_row = sheet.max_row
        max_col = sheet.max_column
        ten_cot = ['0']
        for x in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
            ten_cot.append(x)
        for x in 'ABCDEFGH':
            for y in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
                ten_cot.append(x+y)

        tieu_de = sheet['C3'].value + ' - ' + sheet['C4'].value + ' - ' + sheet['G6'].value + '\n'
        text = tieu_de 
        for i in range(6, max_row + 1):
            line = ''
            for j in range(1, max_col + 1):
                if j not in [5,6]:
                    do_rong_cot = int(sheet.column_dimensions[ten_cot[j]].width) + 1
                    value = sheet.cell(row=i, column=j).value
                    if isinstance(value, str) == False:
                        value = str(value)
                        value = ' '*(do_rong_cot-len(value)) + value
                    else:
                        value = value + ' '*(do_rong_cot-len(value))
                    line = line + value + ' '
            text = text + line + '\n'
        return text

    def get_danh_sach_email(self):
        thong_tin_gv = self.thong_tin.get()
        sheet = self.wb.active
        so_luong_hang = sheet.max_row
        so_luong_cot =  sheet.max_column
        tieu_de = sheet.cell(row=3, column=3).value
        tieu_de = tieu_de + ' - ' + sheet.cell(row=4, column=3).value
        tieu_de = tieu_de + ' - ' + sheet.cell(row=6, column=7).value
        dssv = []
        for i in range(7, so_luong_hang + 1):
            sv = {}
            # Mã SV
            key = sheet.cell(row=6, column=2).value
            value = sheet.cell(row=i, column=2).value
            if isinstance(value, str) == False:
                value = str(value)
            sv[key] = value
            # Họ tên
            key = 'Họ tên'
            value = sheet.cell(row=i, column=3).value + sheet.cell(row=i, column=4).value
            sv[key] = value
            # Điểm
            key = sheet.cell(row=6, column=7).value
            value = sheet.cell(row=i, column=7).value
            if isinstance(value, str) == False:
                value = str(value)
            sv[key] = value
            # Ghi chú
            key = sheet.cell(row=6, column=8).value
            value = sheet.cell(row=i, column=8).value
            if isinstance(value, str) == False:
                value = str(value)
            sv[key] = value

            # Phần của mỗi GV
            for j in range(9, so_luong_cot + 1):
                key = sheet.cell(row=6, column=j).value
                if isinstance(key, str) == False:
                    key = str(key)
                value = sheet.cell(row=i, column=j).value
                if isinstance(value, str) == False:
                    value = str(value)
                sv[key] = value
            dssv.append(sv)
        ds_email = []
        for sv in dssv:
            nguoi = []
            email = sv['Mã SV'] + '@student.hcmute.edu.vn'
            message = 'Subject: ' + tieu_de + '\n\n'
            for key in sv:
                message = message + key + ": " + sv[key] + '\n'
            message =  message + '-'*10 + '\n'
            message = message + '\n' + thong_tin_gv + '\n'
            nguoi.append(email)
            nguoi.append(message)
            ds_email.append(nguoi)
        return ds_email

    def onAbout(self):
        mb.showinfo('About','Trường ĐHSPKT TP.HCM\nKhoa Công nghệ Thông tin\nVersion 2.0 - Tháng 10/2023')

    def onOpenExcel(self):
        self.txt_bang_diem.configure(state = tk.NORMAL)
        self.txt_bang_diem.delete("1.0", tk.END)

        f_types	= [("Excel", "*.xls	*.xlsx")]
        filename = fd.askopenfilename(title="Open Excel File", filetypes=f_types)
        if filename:
            split_tup = os.path.splitext(filename)
            file_extension = split_tup[1]
            if file_extension == '.xls':
                x2x = XLS2XLSX(filename)
                self.wb = x2x.to_xlsx()
            else:
                self.wb = openpyxl.load_workbook(filename)
            danh_sach_sv = self.get_danh_sach_sinh_vien()
            self.ds_email = self.get_danh_sach_email()
            n = len(self.ds_email)
            for i in range(0, n):
                danh_sach_sv = danh_sach_sv + '-'*30 + '\n' + self.ds_email[i][0] + '\n' + self.ds_email[i][1] + '\n' 
            self.txt_bang_diem.insert(tk.INSERT, danh_sach_sv)
            self.txt_bang_diem.configure(state = tk.DISABLED)
            self.lst_ket_qua.delete(0, tk.END)
            self.lst_ket_qua.update()

    def onSendEmail(self):
        self.lst_ket_qua.delete(0, tk.END)
        self.lst_ket_qua.update()
        port = 465  # For SSL
        smtp_server = "smtp.gmail.com"
        sender_email = self.email.get()
        password = self.password.get()
        ds_email = self.ds_email
        n = len(ds_email)
        dem = 1
        for nguoi in ds_email:
            context = ssl.create_default_context()
            with smtplib.SMTP_SSL(smtp_server, port, context=context) as server:
                server.login(sender_email, password)
                receiver_email = nguoi[0]
                message = nguoi[1]
                server.sendmail(sender_email, receiver_email, message.encode('utf-8'))
                s = 'Sent %3d/%d --> %s' % (dem, n, receiver_email)
                self.lst_ket_qua.insert(tk.END, s)
                self.lst_ket_qua.yview_moveto("1.0")
                self.lst_ket_qua.update()
                dem = dem + 1
        s = 'Đã gửi xong %d emails' % n
        mb.showinfo('Thông Báo', s)
        return

if	__name__	==	"__main__":
    app	=	App()
    app.mainloop()
    
